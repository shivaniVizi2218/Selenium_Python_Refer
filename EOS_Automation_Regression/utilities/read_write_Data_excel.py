import openpyxl
import unittest


class Read_Write_Data(unittest.TestCase):

    def get_row_count(self, file, sheetname):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetname)
        return (sheet.max_row)

    def get_column_count(self, file, sheetname):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetname)
        return (sheet.max_column)

    def read_data(self, file, sheetname, rownum, columnnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetname)
        return sheet.cell(row=rownum, column=columnnum).value

    def write_data(self, file, sheetname, rownum, columnnum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetname)
        sheet.cell(row=rownum, column=columnnum).value = data
        workbook.save(file)
