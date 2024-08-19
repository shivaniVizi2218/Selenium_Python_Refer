import sys
from _ast import Assert
import time
import openpyxl
from eospageObjects.bolt_menu import eos_bolt_menu
from utilities.common_util import previous_day_current_date
from utilities.read_write_Data_excel import Read_Write_Data
import os
from selenium.webdriver.common.by import By
from datetime import datetime


class eos_bolt_menu_report(eos_bolt_menu):
    def __init__(self, driver):
        super().__init__(driver)
        self.driver = driver

    akb_tab = "//a[normalize-space()='AKB Stop Notes']"
    attempt_tab = "//a[normalize-space()='Attempt By Facility']"
    previous_day = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])"
    report_iframe = "//div[contains(@class,'viewerContentContainer')]//iframe[1]"
    date_element = "//span[normalize-space()='Report Date']"
    facility_element = "//span[normalize-space()='Facility Name']"
    view_report_button = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@value,'View Report')])"
    facility_click = "(//tr[@id='ParametersRowReportViewerControl']//select[contains(@class,'aspNetDisabled')])"
    facility_selection = "(//tr[@id='ParametersRowReportViewerControl']//select[contains(@class,'aspNetDisabled')])//option[175]"
    report_elements = "//tr[@id='ParametersRowReportViewerControl']//following::tr//span[@aria-label='Report table']//../following::td//descendant::div[2]"
    export_dropdown_menu = "//tr[@id='ParametersRowReportViewerControl']//following::tr//table[@title='Export drop down menu']//img[1]"
    export_dropdown_menu_list = "//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a"
    word_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[1]"
    excel_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[2]"
    power_point_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[3]"
    pdf_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[4]"
    print_button = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//input[@title='Print'])[1]"
    print_dialog_box_text = "(//div[@class='msrs-printdialog-main ui-draggable']//p)[2]"
    dialog_print_button = "(//div[@class='msrs-printdialog-main ui-draggable']//p)[3]"
    Done_button = "(//div[@class='msrs-printdialog-main ui-draggable']//p)[7]"

    alerts_tab = "//a[normalize-space()='Alerts']"

    start_date = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[1]"
    end_date = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[2]"
    dismissed_only = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[3]"
    alert_type = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[4]"
    select_all = "//*[text()='(Select All)']/../input "
    alerts_table_elements = "//tr[@id='ParametersRowReportViewerControl']//following::tr//span[@aria-label='Report table']/../following::tr[5]//td//descendant::div[2]"
    customer = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[2]"
    load_manifest_barcode_elements = "//tr[@id='ParametersRowReportViewerControl']//following::tr//span[@aria-label='Report table']/../following::tr[4]//td//descendant::div[2]"

    def select_akb_tab_reports(self, subtract_days=1):
        self.click_on_element(self.bolt_menu, "id")
        time.sleep(2)
        self.click_on_element(self.bolt_menu_reports, "xpath")
        time.sleep(2)
        self.click_on_element(self.akb_tab, "xpath")
        time.sleep(2)
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        return date_previous_day

    def select_previous_date(self, date_previous_day, expected_texts, expected_list_items, word_file, excel_file_name,
                             power_point_name, pdf_name, text):
        global actual_column_names
        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.date_element, "xpath")
        assert self.isElementPresent(self.facility_element, "xpath")
        assert self.isElementPresent(self.view_report_button, "xpath")
        self.click_on_element(self.previous_day, "xpath")
        self.clear_field(self.previous_day, "xpath")
        self.send_keys_to(self.previous_day, date_previous_day, "xpath")
        self.click_on_element(self.facility_click, "xpath")
        self.click_on_element(self.facility_selection, "xpath")
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(10)
        actual_texts = []
        report_list = self.get_elements(self.report_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            actual_texts.append(option_text)
        assert actual_texts == expected_texts
        time.sleep(5)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_texts_items.append(option_text_items)
        assert actual_texts_items == expected_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        assert word_file == downloaded_word_name
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == excel_file_name
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[2]]
        if actual_column_names == expected_texts:
            assert actual_column_names == expected_texts
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(10)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(10)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")

    def verify_all_the_reports(self, alert_report, drop_down, alert_word, alert_excel, alert_print_text):
        global actual_column_names
        current_date = datetime.now().date()
        formatted_date = current_date.strftime('%m-%d-%Y')
        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        self.click_on_element(self.facility_click, "xpath")
        self.click_on_element(self.facility_selection, "xpath")

        self.click_on_element(self.start_date, "xpath")
        self.clear_field(self.start_date, "xpath")
        self.send_keys_to(self.start_date, "01/01/2022", "xpath")

        self.click_on_element(self.end_date, "xpath")
        self.clear_field(self.end_date, "xpath")
        self.send_keys_to(self.end_date, formatted_date, "xpath")

        self.click_on_element(self.dismissed_only, "xpath")
        self.click_on_element(self.select_all, "xpath")
        self.click_on_element(self.alert_type, "xpath")
        self.click_on_element(self.select_all, "xpath")
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(20)
        actual_texts = []
        alert_report_list = self.get_elements(self.alerts_table_elements, "xpath")
        for option in alert_report_list:
            option_text = option.get_attribute('textContent')
            actual_texts.append(option_text)
        assert actual_texts == alert_report
        time.sleep(5)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_texts_items.append(option_text_items)
        assert actual_texts_items == drop_down
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(35)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        assert alert_word == downloaded_word_name
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(35)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == alert_excel
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[2]]
        if actual_column_names == alert_report:
            assert actual_column_names == alert_report
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert alert_print_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(15)
        self.click_on_element(self.Done_button, "xpath")

    def verify_attempt_facility_reports(self, attempt_expected_texts, expected_attempt_list_items, attempt_word_file,
                                        attempt_excel_file_name, attempt_power_point_name, attempt_pdf_name,
                                        attempt_text, subtract_days=1):
        global actual_column_names
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)

        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.end_date, "xpath")
        assert self.isElementPresent(self.facility_click, "xpath")
        assert self.isElementPresent(self.dismissed_only, "xpath")
        assert self.isElementPresent(self.view_report_button, "xpath")
        self.click_on_element(self.end_date, "xpath")
        self.clear_field(self.end_date, "xpath")
        self.send_keys_to(self.end_date, date_previous_day, "xpath")
        self.click_on_element(self.facility_click, "xpath")
        self.click_on_element(self.facility_selection, "xpath")
        time.sleep(3)
        self.click_on_element(self.dismissed_only, "xpath")
        self.click_on_element(self.select_all, "xpath")
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(10)
        attempt_actual_texts = []
        report_list = self.get_elements(self.report_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            attempt_actual_texts.append(option_text)
        assert attempt_actual_texts == attempt_expected_texts
        time.sleep(5)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_attempt_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_attempt_texts_items.append(option_text_items)
        assert actual_attempt_texts_items == expected_attempt_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        assert attempt_word_file == downloaded_word_name
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == attempt_excel_file_name
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[2]]
        if actual_column_names == expected_attempt_list_items:
            assert actual_column_names == expected_attempt_list_items
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(30)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == attempt_power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(30)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == attempt_pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert attempt_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")

    def verify_driver_summary_reports(self, driver_info, expected_diver_list_items, driver_word_file, driver_excel,
                                      driver_power_point_name, driver_pdf_name, driver_text):
        global actual_column_names
        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.start_date, "xpath")
        assert self.isElementPresent(self.facility_click, "xpath")
        assert self.isElementPresent(self.customer, "xpath")
        assert self.isElementPresent(self.dismissed_only, "xpath")

        self.click_on_element(self.start_date, "xpath")
        self.clear_field(self.start_date, "xpath")
        self.send_keys_to(self.start_date, "02/22/2022", "xpath")
        time.sleep(10)
        self.click_on_element(self.facility_click, "xpath")
        time.sleep(10)
        self.click_on_element(self.facility_click, "xpath")
        time.sleep(10)
        self.click_on_element(self.facility_selection, "xpath")
        time.sleep(30)
        self.click_on_element(self.customer, "xpath")
        time.sleep(10)
        self.click_on_element(self.select_all, "xpath")
        time.sleep(10)
        self.click_on_element(self.start_date, "xpath")
        time.sleep(30)
        self.click_on_element(self.dismissed_only, "xpath")
        time.sleep(10)
        self.click_on_element(self.select_all, "xpath")
        time.sleep(10)
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(30)
        driver_actual_texts = []
        report_list = self.get_elements(self.alerts_table_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            driver_actual_texts.append(option_text)
        print(driver_actual_texts)
        assert driver_info == driver_actual_texts
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_attempt_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_attempt_texts_items.append(option_text_items)
        assert actual_attempt_texts_items == expected_diver_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert driver_word_file == downloaded_word_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == driver_excel
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[2]]
        if actual_column_names == driver_info:
            assert actual_column_names == driver_info
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(30)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == driver_power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(30)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == driver_pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert driver_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")

    def verify_load_manifest_reports(self, manifest_info, manifest_list_items, manifest_word_file, manifest_excel,
                                     manifest_power_point_name, manifest_pdf_name, manifest_text, subtract_days=1):
        global actual_column_names
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.start_date, "xpath")
        assert self.isElementPresent(self.facility_click, "xpath")
        assert self.isElementPresent(self.customer, "xpath")
        self.click_on_element(self.start_date, "xpath")
        self.clear_field(self.start_date, "xpath")
        self.send_keys_to(self.start_date, date_previous_day, "xpath")
        self.click_on_element(self.facility_click, "xpath")
        time.sleep(3)
        self.click_on_element(self.facility_selection, "xpath")
        time.sleep(15)
        self.click_on_element(self.customer, "xpath")
        time.sleep(5)
        self.click_on_element(self.select_all, "xpath")
        time.sleep(2)
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(20)
        manifest_actual_texts = []
        report_list = self.get_elements(self.load_manifest_barcode_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            manifest_actual_texts.append(option_text)
        assert manifest_info == manifest_actual_texts
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_attempt_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_attempt_texts_items.append(option_text_items)
        assert actual_attempt_texts_items == manifest_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert manifest_word_file == downloaded_word_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == manifest_excel
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[2]]
        if actual_column_names == manifest_info:
            assert actual_column_names == manifest_info
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(30)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == manifest_power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(30)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == manifest_pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert manifest_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")

    def load_manifest_stop_reports(self, manifest_stop_info, manifest_stop_list_items, manifest_stop_word_file,
                                   manifest_stop_excel,
                                   manifest_stop_power_point_name, manifest_stop_pdf_name, manifest_stop_text,
                                   subtract_days=1):
        global actual_column_names
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.start_date, "xpath")
        assert self.isElementPresent(self.facility_click, "xpath")
        assert self.isElementPresent(self.customer, "xpath")
        self.click_on_element(self.start_date, "xpath")
        self.clear_field(self.start_date, "xpath")
        self.send_keys_to(self.start_date, date_previous_day, "xpath")
        self.click_on_element(self.facility_click, "xpath")
        time.sleep(3)
        self.click_on_element(self.facility_selection, "xpath")
        time.sleep(30)
        self.click_on_element(self.customer, "xpath")
        time.sleep(5)
        self.click_on_element(self.select_all, "xpath")
        time.sleep(2)
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(20)
        manifest_stop_actual_texts = []
        report_list = self.get_elements(self.load_manifest_barcode_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            manifest_stop_actual_texts.append(option_text)
        assert manifest_stop_info == manifest_stop_actual_texts
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_attempt_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_attempt_texts_items.append(option_text_items)
        assert actual_attempt_texts_items == manifest_stop_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert manifest_stop_word_file == downloaded_word_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == manifest_stop_excel
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[2]]
        if actual_column_names == manifest_stop_info:
            assert actual_column_names == manifest_stop_info
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(30)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == manifest_stop_power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(30)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == manifest_stop_pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert manifest_stop_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")

    def load_manifest_received_reports(self, manifest_received_info, manifest_received_list_items,
                                       manifest_received_word_file, manifest_received_excel,
                                       manifest_received_power_point_name, manifest_received_pdf_name,
                                       manifest_received_text, subtract_days=1):
        global actual_column_names
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.start_date, "xpath")
        assert self.isElementPresent(self.facility_click, "xpath")
        assert self.isElementPresent(self.customer, "xpath")
        self.click_on_element(self.start_date, "xpath")
        self.clear_field(self.start_date, "xpath")
        self.send_keys_to(self.start_date, date_previous_day, "xpath")
        self.click_on_element(self.facility_click, "xpath")
        time.sleep(3)
        self.click_on_element(self.facility_selection, "xpath")
        time.sleep(30)
        self.click_on_element(self.customer, "xpath")
        time.sleep(5)
        self.click_on_element(self.select_all, "xpath")
        time.sleep(2)
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(20)
        manifest_received_actual_texts = []
        report_list = self.get_elements(self.load_manifest_barcode_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            manifest_received_actual_texts.append(option_text)
        assert manifest_received_info == manifest_received_actual_texts
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_attempt_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_attempt_texts_items.append(option_text_items)
        assert actual_attempt_texts_items == manifest_received_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert manifest_received_word_file == downloaded_word_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == manifest_received_excel
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[2]]
        if actual_column_names == manifest_received_info:
            assert actual_column_names == manifest_received_info
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(30)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == manifest_received_power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(30)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == manifest_received_pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert manifest_received_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")

    def verify_all_services_reports(self,location_services_info, location_services_list_items, location_services_word_file,
                                    location_services_excel, location_services_power_point_name, location_services_pdf_name, location_services_text, subtract_days=1):
        global actual_column_names
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.switch_to.frame(self.get_element(self.report_iframe, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.start_date, "xpath")
        assert self.isElementPresent(self.customer, "xpath")
        self.click_on_element(self.start_date, "xpath")
        self.clear_field(self.start_date, "xpath")
        self.send_keys_to(self.start_date, date_previous_day, "xpath")
        time.sleep(10)
        self.click_on_element(self.customer, "xpath")
        time.sleep(5)
        self.click_on_element(self.select_all, "xpath")
        time.sleep(4)
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(20)
        location_services_actual_texts = []
        report_list = self.get_elements(self.report_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            location_services_actual_texts.append(option_text)
        assert location_services_info == location_services_actual_texts
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        actual_attempt_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            actual_attempt_texts_items.append(option_text_items)
        assert actual_attempt_texts_items == location_services_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert location_services_word_file == downloaded_word_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == location_services_excel
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[1]]
        if actual_column_names == location_services_info:
            assert actual_column_names == location_services_info
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(30)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == location_services_power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(30)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == location_services_pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert location_services_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")
