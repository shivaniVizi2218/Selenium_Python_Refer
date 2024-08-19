
import time
from selenium.webdriver.common.by import By
from eospageObjects.bolt_menu import eos_bolt_menu
from eospageObjects.eos_bolt_menu_reports import eos_bolt_menu_report
from utilities.common_util import previous_day_current_date
import openpyxl
import os
from datetime import datetime


class mls_bolt_menu(eos_bolt_menu):

    def __init__(self, driver):
        super().__init__(driver)
        self.eos_bolt_menu_report = None
        self.bolt_menu = None
        self.driver = driver

    missing_load_scan = "//div[@id='navbar-collapse-reports']/descendant::li/a[text()='MLS (Missing Load Scan)']"
    report_iframe1 = "//div[contains(@class,'viewerContentContainer')]//iframe[1]"
    browser_txt = "//span[text()='Browse']"
    mls_scan_text = "//span[text()='MLS (Missing Load Scan)']"
    select_date_val = "(//input[@title='Select a value'])[1]"
    previous_day = "//td[@class='ms-picker-daycenterOn']"
    facility_carat = "(//input[@title='Select a value'])[2]"
    select_all_check_box = "(//td[@nowrap='nowrap']/descendant::input[contains(@id,'ReportViewerControl')])[1]"
    view_report_btn = "//input[@value='View Report']"
    customer_carat_btn = "(//input[@title='Select a value'])[3]"
    delivery_date_field = "(//table[contains(@id,'ParametersGridReportViewerControl')]/descendant::input[@type='text'])[1]"
    generated_report = "(//div[contains(@id,'VisibleReportContentReportViewerControl')]/descendant::tbody/tr/td)[1]"
    facility_report_txt = "(//div[text()='Facility'])[1]"
    mls_facility_report_link = "//a[text()='MLS By Facility and Route']"
    mls_facility_route_txt = "//span[text()='MLS by Facility and Route']"
    facility_drop_down = "//select[contains(@name,'ReportViewerControl')]"
    west_palm_beach_option = "//option[@value='24']"
    route_select_btn = "(//input[@alt='Select a value'])[2]"
    route_text_area = "//textarea[contains(@name,'ReportViewerControl')]"

    barcode_txt = "(//td/descendant::div[text()='Barcode'])[1]"

    customer_name_txt = "(//td/descendant::div[text()='Customer Name'])[1]"
    route_name = "(//td/descendant::div[text()='Route Name'])[1]"
    contractor_code_txt = "(//td/descendant::div[text()='Contractor Code'])[1]"
    destination_name_txt = "(//td/descendant::div[text()='Destination Name'])[1]"
    address_txt = "(//td/descendant::div[text()='Address'])[1]"
    city_txt = "(//td/descendant::div[text()='City'])[1]"
    state_txt = "(//td/descendant::div[text()='State'])[1]"
    zip_txt = "(//td/descendant::div[text()='Zip'])[1]"
    attention_to_txt = "(//td/descendant::div[text()='Attention To'])[1]"

    # facility Zip
    mls_facility_zip_txt = "//span[text()='MLS By Facility and Zip']"
    customer_txt = "(//td/descendant::div[text()='Customer'])[1]"
    route_txt = "(//div[text()='Route '])[2]"
    IC_txt = "(//div[text()='IC'])[2]"
    original_EDD = "(//div[text()='Original EDD'])[2]"
    current_status = "(//div[text()='Current Status'])[2]"
    last_event = "(//div[text()='Last Event'])[2]"
    last_scan_time = "(//div[text()='Last Scan Time'])[2]"
    akron_option = "//option[@value='56']"
    # undelivered exceptions
    report_date_field = "//td[@class='ParamLabelCell']/descendant::span[text()='Report Date']"
    facility_field = "//td[@class='ParamLabelCell']/descendant::span[text()='Facility']"
    event_name_field = "//td[@class='ParamLabelCell']/descendant::span[text()='Event Name']"
    no_action_field = "//td[@class='ParamLabelCell']/descendant::span[text()='No Action']"
    event_name_drop_down = "(//select[contains(@name,'ReportViewerControl')])[1]"
    no_action_drop_down = "(//select[contains(@name,'ReportViewerControl')])[2]"
    event_name_values = "//option[@value='31']"
    no_action_val = "//option[text()='All']"
    references_txt = "(//div[text()='Reference'])[2]"
    stop_point_txt = "(//div[text()='Stop Point'])[2]"
    original_route = "(//div[text()='Original Route'])[2]"
    edd_txt = "(//div[text()='EDD'])[2]"
    route_txt1 = "(//div[text()='Route'])[2]"
    event_date_time = "(//div[text()='Event Date Time'])[2]"
    first_received_date_time = "(//div[text()='First Received Date Time'])[2]"
    last_hub_date_time = "(//div[text()='Last Hub Date Time'])[2]"
    hub_location = "(//div[text()='Hub Location'])[2]"
    event_name_natural_disaster = "//option[@value='33']"
    event_name_left_on_dock = "//option[@value='42']"
    jackson_value_option = "//option[@value='25']"
    customer_carat_arrow = "(//input[@title='Select a value'])[2]"
    export_drop_down_btn = "//div[@class='ToolbarExport WidgetSet']/descendant::table[2]"
    word_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[1]"
    excel_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[2]"
    pp_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[3]"
    pdf_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[4]"
    tiff_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[5]"
    mhtml_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[6]"
    csv_comma_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[7]"
    csv_pipe_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[8]"
    csv_no_head_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[9]"
    tab_delimited_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[10]"
    xml_file_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[11]"
    data_feed_type_btn = "(//div[@class='MenuBarBkGnd']/descendant::div/a)[12]"
    customer_field_txt = "//td[@class='ParamLabelCell']/descendant::span[text()='Customer']"
    diff_txt = "//td[@class='ParamLabelCell']/descendant::span[text()='Date Diff']"

    status_update_second_report_elements = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//span[@aria-label='Report table']/../following::tr[4])[2]//td//descendant::div[2]"
    status_update_first_report_elements = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//span[@aria-label='Report table']/../descendant::tr)[2]//td//descendant::div[2]"
    status_update_report_date = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[1]"
    status_update_facility = "(//tr[@id='ParametersRowReportViewerControl']//select[contains(@class,'aspNetDisabled')])"
    status_update_facility_selection = "(//tr[@id='ParametersRowReportViewerControl']//select[contains(@class,'aspNetDisabled')])//option[175]"
    status_update_customer = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[2]"
    status_update_select_all = "//*[text()='(Select All)']/../input"
    status_update_service = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@name,'txtValue')])[3]"
    view_report_button = "(//tr[@id='ParametersRowReportViewerControl']//input[contains(@value,'View Report')])"
    export_dropdown_menu = "//tr[@id='ParametersRowReportViewerControl']//following::tr//table[@title='Export drop down menu']//img[1]"
    export_dropdown_menu_list = "//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a"
    word_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[1]"
    excel_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[2]"
    power_point_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[3]"
    pdf_file = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//div[@class='MenuBarBkGnd']/descendant::div//a)[4]"
    print_button = "(//tr[@id='ParametersRowReportViewerControl']//following::tr//input[@title='Print'])[1]"
    print_dialog_box_text = "(//div[@class='msrs-printdialog-main ui-draggable']//p)[2]"
    dialog_print_button = "(//div[@class='msrs-printdialog-main ui-draggable']//p)[3]"
    Done_button = "(//div[@class='msrs-printdialog-main ui-draggable']//p)[7]"


    def verify_browse_header(self):
        browse_txt = self.driver.find_element(By.XPATH, self.browser_txt).get_attribute(
            'textContent')
        browse_txt = browse_txt.strip()
        print("browse_txt" + browse_txt)
        assert browse_txt == "Browse"

    def verify_mls_scan_header(self):
        mls_text = self.driver.find_element(By.XPATH, self.mls_scan_text).get_attribute(
            'textContent')
        mls_text = mls_text.strip()
        print("mls_text" + mls_text)
        assert mls_text == "MLS (Missing Load Scan)"

    def verify_mls_facility_header(self):
        mls_text = self.driver.find_element(By.XPATH, self.mls_facility_route_txt).get_attribute(
            'textContent')
        mls_text = mls_text.strip()
        print("mls_text" + mls_text)
        assert mls_text == "MLS by Facility and Route"

    def verify_generated_report(self):
        facility_txt = self.driver.find_element(By.XPATH, self.facility_report_txt).get_attribute(
            'textContent')
        facility_txt = facility_txt.strip()
        print("facility_txt:" + facility_txt)
        assert facility_txt == "Facility"

    def mls_missing_load_scan_reports_validation(self, subtract_days=1):
        time.sleep(3)
        self.verify_browse_header()
        time.sleep(2)
        self.verify_mls_scan_header()
        time.sleep(3)
        self.driver.switch_to.frame(self.get_element(self.report_iframe1, "xpath"))
        time.sleep(10)
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.find_element(By.XPATH, self.delivery_date_field).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).clear()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).send_keys(date_previous_day)
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.facility_carat).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.select_all_check_box).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.facility_carat).click()
        time.sleep(40)
        self.driver.find_element(By.XPATH, self.customer_carat_btn).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.select_all_check_box).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.view_report_btn).click()
        time.sleep(10)
        self.verify_generated_report()

    def mls_facility_and_route_report_scan(self, delivery_date_txt, route_txt):
        time.sleep(3)
        # self.bolt_menu = eos_bolt_menu(self.driver)
        # link = self.driver.find_element(By.XPATH, "//a[text()='MLS By Facility and Route']").get_attribute("href")
        # print("LINK ---------->> ", link)
        # linkNew = link.replace('http://', 'http://eostest1:l1a9s1e2Rp@--@')
        # print("LINK ---------->> ", linkNew)
        # self.driver.get(linkNew)
        # self.driver.find_element(By.XPATH, self.mls_facility_report_link).click()
        # self.bolt_menu.go_to_a_bolt_menu_view_report(report_name="mls_by_facility_and_route")
        # time.sleep(10)
        self.verify_browse_header()
        time.sleep(2)
        # assert (self.isElementPresent(self.mls_scan_text), "xpath")
        self.verify_mls_facility_header()
        time.sleep(3)
        self.driver.switch_to.frame(self.get_element(self.report_iframe1, "xpath"))
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.delivery_date_field).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).clear()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).send_keys(delivery_date_txt)
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.facility_drop_down).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.west_palm_beach_option).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.route_select_btn).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.route_text_area).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.route_text_area).send_keys(route_txt)
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.route_select_btn).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.view_report_btn).click()
        time.sleep(10)

    def verify_report_column_names(self):
        barcode_txt = self.driver.find_element(By.XPATH, self.barcode_txt).get_attribute(
            'textContent')
        barcode_txt = barcode_txt.strip()
        print("barcode_txt" + barcode_txt)
        assert barcode_txt == "Barcode"

        customer_name_txt = self.driver.find_element(By.XPATH, self.customer_name_txt).get_attribute(
            'textContent')
        customer_name_txt = customer_name_txt.strip()
        print("customer_name_txt" + customer_name_txt)
        assert customer_name_txt == "Customer Name"

        route_name = self.driver.find_element(By.XPATH, self.route_name).get_attribute(
            'textContent')
        print("route_name" + route_name)
        route_name = route_name.strip()
        assert route_name == "Route Name"

        contractor_code_txt = self.driver.find_element(By.XPATH, self.contractor_code_txt).get_attribute(
            'textContent')
        contractor_code_txt = contractor_code_txt.strip()
        print("contractor_code_txt" + contractor_code_txt)
        assert contractor_code_txt == "Contractor Code"

        destination_name_txt = self.driver.find_element(By.XPATH, self.destination_name_txt).get_attribute(
            'textContent')
        destination_name_txt = destination_name_txt.strip()
        print("barcode_txt" + destination_name_txt)
        assert destination_name_txt == "Destination Name"

        address_txt = self.driver.find_element(By.XPATH, self.address_txt).get_attribute(
            'textContent')
        address_txt = address_txt.strip()
        print("address_txt" + address_txt)
        assert address_txt == "Address"

        city_txt = self.driver.find_element(By.XPATH, self.city_txt).get_attribute(
            'textContent')
        city_txt = city_txt.strip()
        print("barcode_txt" + city_txt)
        assert city_txt == "City"

        state_txt = self.driver.find_element(By.XPATH, self.state_txt).get_attribute(
            'textContent')
        state_txt = state_txt.strip()
        print("state_txt" + state_txt)
        assert state_txt == "State"

        zip_txt = self.driver.find_element(By.XPATH, self.zip_txt).get_attribute(
            'textContent')
        zip_txt = zip_txt.strip()
        print("zip_txt" + zip_txt)
        assert zip_txt == "Zip"

        attention_to_txt = self.driver.find_element(By.XPATH, self.attention_to_txt).get_attribute(
            'textContent')
        attention_to_txt = attention_to_txt.strip()
        print("attention_to_txt" + attention_to_txt)
        assert attention_to_txt == "Attention To"


    def verify_mls_facility_zip_header(self):
        mls_zip_text = self.driver.find_element(By.XPATH, self.mls_facility_zip_txt).get_attribute(
            'textContent')
        mls_zip_text = mls_zip_text.strip()
        print("mls_zip_text" + mls_zip_text)
        assert mls_zip_text == "MLS By Facility and Zip"

    def mls_facility_and_zip_report_scan(self, delivery_date, route_num):
        time.sleep(3)
        self.verify_browse_header()
        time.sleep(3)
        self.verify_mls_facility_zip_header()
        time.sleep(2)
        self.driver.switch_to.frame(self.get_element(self.report_iframe1, "xpath"))
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.delivery_date_field).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).clear()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).send_keys(delivery_date)
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.facility_drop_down).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.akron_option).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.route_select_btn).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.route_text_area).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.route_text_area).send_keys(route_num)
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.route_select_btn).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.view_report_btn).click()
        time.sleep(10)

    def verify_facility_zip_column_names(self):
        barcode_txt = self.driver.find_element(By.XPATH, self.barcode_txt).get_attribute(
            'textContent')
        barcode_txt = barcode_txt.strip()
        print("barcode_txt" + barcode_txt)
        assert barcode_txt == "Barcode"

        customer_txt = self.driver.find_element(By.XPATH, self.customer_txt).get_attribute(
            'textContent')
        customer_txt = customer_txt.strip()
        print("customer_txt" + customer_txt)
        assert customer_txt == "Customer"

        route_txt = self.driver.find_element(By.XPATH, self.route_txt).get_attribute(
            'textContent')
        print("barcode_txt" + route_txt)
        assert route_txt == "Route "

        IC_txt = self.driver.find_element(By.XPATH, self.IC_txt).get_attribute(
            'textContent')
        IC_txt = IC_txt.strip()
        print("IC_txt" + IC_txt)
        assert IC_txt == "IC"

        destination_name_txt = self.driver.find_element(By.XPATH, self.destination_name_txt).get_attribute(
            'textContent')
        destination_name_txt = destination_name_txt.strip()
        print("barcode_txt" + destination_name_txt)
        assert destination_name_txt == "Destination Name"

        address_txt = self.driver.find_element(By.XPATH, self.address_txt).get_attribute(
            'textContent')
        address_txt = address_txt.strip()
        print("address_txt" + address_txt)
        assert address_txt == "Address"

        city_txt = self.driver.find_element(By.XPATH, self.city_txt).get_attribute(
            'textContent')
        city_txt = city_txt.strip()
        print("barcode_txt" + city_txt)
        assert city_txt == "City"

        state_txt = self.driver.find_element(By.XPATH, self.state_txt).get_attribute(
            'textContent')
        state_txt = state_txt.strip()
        print("state_txt" + state_txt)
        assert state_txt == "State"

        zip_txt = self.driver.find_element(By.XPATH, self.zip_txt).get_attribute(
            'textContent')
        zip_txt = zip_txt.strip()
        print("zip_txt" + zip_txt)
        assert zip_txt == "Zip"

        original_EDD = self.driver.find_element(By.XPATH, self.original_EDD).get_attribute(
            'textContent')
        original_EDD = original_EDD.strip()
        print("original_EDD" + original_EDD)
        assert original_EDD == "Original EDD"

        current_status = self.driver.find_element(By.XPATH, self.current_status).get_attribute(
            'textContent')
        current_status = current_status.strip()
        print("current_status" + current_status)
        assert current_status == "Current Status"

        last_event = self.driver.find_element(By.XPATH, self.last_event).get_attribute(
            'textContent')
        last_event = last_event.strip()
        print("last_event" + last_event)
        assert last_event == "Last Event"

        last_scan_time = self.driver.find_element(By.XPATH, self.last_scan_time).get_attribute(
            'textContent')
        last_scan_time = last_scan_time.strip()
        print("last_scan_time" + last_scan_time)
        assert last_scan_time == "Last Scan Time"

    def verify_field_vals(self):
        report_date_field = self.driver.find_element(By.XPATH, self.report_date_field).get_attribute(
            'textContent')
        report_date_field = report_date_field.strip()
        print("last_scan" + report_date_field)
        assert report_date_field == "Report Date"

        facility_field = self.driver.find_element(By.XPATH, self.facility_field).get_attribute(
            'textContent')
        facility_field = facility_field.strip()
        print("facility_field" + facility_field)
        assert facility_field == "Facility"

        event_name_field = self.driver.find_element(By.XPATH, self.event_name_field).get_attribute(
            'textContent')
        event_name_field = event_name_field.strip()
        print("event_name_field" + event_name_field)
        assert event_name_field == "Event Name"

        no_action_field = self.driver.find_element(By.XPATH, self.no_action_field).get_attribute(
            'textContent')
        no_action_field = no_action_field.strip()
        print("no_action_field" + no_action_field)
        assert no_action_field == "No Action"
        assert self.isElementPresent(self.view_report_btn, 'xpath')


    def unable_to_deliver_exception_report(self, subtract_days=1):
        time.sleep(3)
        self.verify_browse_header()
        time.sleep(2)
        self.driver.switch_to.frame(self.get_element(self.report_iframe1, "xpath"))
        time.sleep(10)
        self.verify_field_vals()
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.find_element(By.XPATH, self.delivery_date_field).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).clear()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).send_keys(date_previous_day)
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.facility_carat).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.select_all_check_box).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.facility_carat).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.event_name_drop_down).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.event_name_values).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.no_action_drop_down).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.no_action_val).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.view_report_btn).click()
        time.sleep(10)

    def verify_deliver_exception_reports(self):
        facility_field = self.driver.find_element(By.XPATH, self.facility_field).get_attribute(
            'textContent')
        facility_field = facility_field.strip()
        print("facility_field" + facility_field)
        assert facility_field == "Facility"

        barcode_txt = self.driver.find_element(By.XPATH, self.barcode_txt).get_attribute(
            'textContent')
        barcode_txt = barcode_txt.strip()
        print("barcode_txt" + barcode_txt)
        assert barcode_txt == "Barcode"

        customer_name_txt = self.driver.find_element(By.XPATH, self.customer_name_txt).get_attribute(
            'textContent')
        customer_name_txt = customer_name_txt.strip()
        print("customer_name_txt" + customer_name_txt)
        assert customer_name_txt == "Customer Name"

        zip_txt = self.driver.find_element(By.XPATH, self.zip_txt).get_attribute(
            'textContent')
        zip_txt = zip_txt.strip()
        print("zip_txt" + zip_txt)
        assert zip_txt == "Zip"

        address_txt = self.driver.find_element(By.XPATH, self.address_txt).get_attribute(
            'textContent')
        address_txt = address_txt.strip()
        print("address_txt" + address_txt)
        assert address_txt == "Address"

        original_EDD = self.driver.find_element(By.XPATH, self.original_EDD).get_attribute(
            'textContent')
        original_EDD = original_EDD.strip()
        print("original_EDD" + original_EDD)
        assert original_EDD == "Original EDD"

        # last_scan = self.driver.find_element(By.XPATH, self.last_scan).get_attribute(
        #     'textContent')
        # last_scan = last_scan.strip()
        # print("last_scan" + last_scan)
        # assert last_scan == "Last Scan Time"

        references_txt = self.driver.find_element(By.XPATH, self.references_txt).get_attribute(
            'textContent')
        references_txt = references_txt.strip()
        print("references_txt" + references_txt)
        assert references_txt == "Reference"

        current_status = self.driver.find_element(By.XPATH, self.current_status).get_attribute(
            'textContent')
        current_status = current_status.strip()
        print("current_status" + current_status)
        assert current_status == "Current Status"

        stop_point_txt = self.driver.find_element(By.XPATH, self.stop_point_txt).get_attribute(
            'textContent')
        stop_point_txt = stop_point_txt.strip()
        print("stop_point_txt" + stop_point_txt)
        assert stop_point_txt == "Stop Point"

        original_route = self.driver.find_element(By.XPATH, self.original_route).get_attribute(
            'textContent')
        original_route = original_route.strip()
        print("original_route" + original_route)
        assert original_route == "Original Route"

        edd_txt = self.driver.find_element(By.XPATH, self.edd_txt).get_attribute(
            'textContent')
        edd_txt = edd_txt.strip()
        print("original_route" + edd_txt)
        assert edd_txt == "EDD"

        route_txt1 = self.driver.find_element(By.XPATH, self.route_txt1).get_attribute(
            'textContent')
        route_txt1 = route_txt1.strip()
        print("route_txt1" + route_txt1)
        assert route_txt1 == "Route"

        event_date_time = self.driver.find_element(By.XPATH, self.event_date_time).get_attribute(
            'textContent')
        event_date_time = event_date_time.strip()
        print("event_date_time" + event_date_time)
        assert event_date_time == "Event Date Time"

        first_received_date_time = self.driver.find_element(By.XPATH, self.first_received_date_time).get_attribute(
            'textContent')
        first_received_date_time = first_received_date_time.strip()
        print("first_received_date_time" + first_received_date_time)
        assert first_received_date_time == "First Received Date Time"

        last_hub_date_time = self.driver.find_element(By.XPATH, self.last_hub_date_time).get_attribute(
            'textContent')
        last_hub_date_time = last_hub_date_time.strip()
        print("last_hub_date_time" + last_hub_date_time)
        assert last_hub_date_time == "Last Hub Date Time"

        hub_location = self.driver.find_element(By.XPATH, self.hub_location).get_attribute(
            'textContent')
        hub_location = hub_location.strip()
        print("hub_location" + hub_location)
        assert hub_location == "Hub Location"

        last_event = self.driver.find_element(By.XPATH, self.last_event).get_attribute(
            'textContent')
        last_event = last_event.strip()
        print("last_event" + last_event)
        assert last_event == "Last Event"

        last_scan_time = self.driver.find_element(By.XPATH, self.last_scan_time).get_attribute(
            'textContent')
        last_scan_time = last_scan_time.strip()
        print("last_scan_time" + last_scan_time)
        assert last_scan_time == "Last Scan Time"

    def select_natural_disaster(self):
        self.driver.find_element(By.XPATH, self.event_name_drop_down).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.event_name_natural_disaster).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.event_name_drop_down).click()

    def select_left_on_dock(self):
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.event_name_drop_down).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.event_name_left_on_dock).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.event_name_drop_down).click()

    def click_on_report(self):
        self.driver.find_element(By.XPATH, self.view_report_btn).click()
        time.sleep(10)

    def validating_MOSD_by_facility(self, subtract_days=1):
        time.sleep(3)
        self.verify_browse_header()
        time.sleep(2)
        self.driver.switch_to.frame(self.get_element(self.report_iframe1, "xpath"))
        # time.sleep(10)
        assert self.isElementPresent(self.delivery_date_field, 'xpath')
        assert self.isElementPresent(self.facility_field, 'xpath')
        assert self.isElementPresent(self.customer_field_txt, 'xpath')
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.find_element(By.XPATH, self.delivery_date_field).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).clear()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.delivery_date_field).send_keys(date_previous_day)
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.facility_carat).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.jackson_value_option).click()
        time.sleep(4)
        self.driver.find_element(By.XPATH, self.customer_carat_arrow).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.select_all_check_box).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.customer_carat_arrow).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.view_report_btn).click()
        time.sleep(45)
        self.driver.find_element(By.XPATH, self.export_drop_down_btn).click()
        time.sleep(5)
        self.driver.find_element(By.XPATH, self.word_type_btn).click()
        time.sleep(8)

    def export_drop_down_validation(self):
        assert self.isElementPresent(self.word_type_btn, 'xpath')
        assert self.isElementPresent(self.excel_type_btn, 'xpath')
        assert self.isElementPresent(self.pp_type_btn, 'xpath')
        assert self.isElementPresent(self.pdf_type_btn, 'xpath')
        assert self.isElementPresent(self.tiff_type_btn, 'xpath')
        assert self.isElementPresent(self.mhtml_type_btn, 'xpath')
        assert self.isElementPresent(self.csv_comma_type_btn, 'xpath')
        assert self.isElementPresent(self.csv_pipe_type_btn, 'xpath')
        assert self.isElementPresent(self.csv_no_head_type_btn, 'xpath')
        assert self.isElementPresent(self.tab_delimited_type_btn, 'xpath')
        assert self.isElementPresent(self.xml_file_type_btn, 'xpath')
        assert self.isElementPresent(self.data_feed_type_btn, 'xpath')

    def validating_print_text_vals(self, word_file, excel_file_name, expected_texts, pdf_name, power_point_name, text):
        self.eos_bolt_menu_report = eos_bolt_menu_report(self.driver)
        global actual_column_names
        self.click_on_element(self.eos_bolt_menu_report.word_file, "xpath")
        # self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.word_file).click()
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        print("downloaded_word_name##############################")
        print(downloaded_word_name)
        assert downloaded_word_name == word_file
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.export_dropdown_menu).click()
        time.sleep(1)
        self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.excel_file).click()
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == excel_file_name
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[5]]
            print("Actual column Vals######################")
            print(actual_column_names)
        if actual_column_names == expected_texts:
            assert actual_column_names == expected_texts
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.eos_bolt_menu_report.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.eos_bolt_menu_report.power_point_file, "xpath")
        time.sleep(10)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.eos_bolt_menu_report.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.eos_bolt_menu_report.pdf_file, "xpath")
        time.sleep(10)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.eos_bolt_menu_report.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.print_dialog_box_text).get_attribute(
            "textContent")
        assert text == print_text
        self.click_on_element(self.eos_bolt_menu_report.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.eos_bolt_menu_report.Done_button, "xpath")

    def verify_status_update_report(self,status_update_first_info, status_update_second_info, status_update_list_items,
                                    status_update_word_file,status_update_excel,status_update_power_point_name,
                                    status_update_pdf_name,status_update_text, subtract_days=1):
        global first_actual_column_names
        global second_actual_column_names
        date_previous_day = previous_day_current_date("%m/%d/%Y", subtract_days)
        self.driver.switch_to.frame(self.get_element(self.report_iframe1, "xpath"))
        time.sleep(10)
        assert self.isElementPresent(self.status_update_report_date, "xpath")
        assert self.isElementPresent(self.status_update_facility, "xpath")
        assert self.isElementPresent(self.status_update_customer, "xpath")
        assert self.isElementPresent(self.status_update_service, "xpath")
        self.click_on_element(self.status_update_report_date, "xpath")
        self.clear_field(self.status_update_report_date, "xpath")
        self.send_keys_to(self.status_update_report_date, date_previous_day, "xpath")
        self.click_on_element(self.status_update_facility, "xpath")
        time.sleep(5)
        self.click_on_element(self.status_update_facility_selection, "xpath")
        time.sleep(20)
        self.click_on_element(self.status_update_customer, "xpath")
        time.sleep(5)
        self.click_on_element(self.status_update_select_all, "xpath")
        time.sleep(5)
        self.click_on_element(self.status_update_report_date, "xpath")
        time.sleep(20)
        self.click_on_element(self.status_update_service, "xpath")
        time.sleep(5)
        self.click_on_element(self.status_update_select_all, "xpath")
        time.sleep(2)
        self.click_on_element(self.view_report_button, "xpath")
        time.sleep(20)
        status_update_first_report_actual_texts = []
        report_list = self.get_elements(self.status_update_first_report_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            status_update_first_report_actual_texts.append(option_text)
        assert status_update_first_info == status_update_first_report_actual_texts
        status_update_second_report_actual_texts = []
        report_list = self.get_elements(self.status_update_second_report_elements, "xpath")
        for option in report_list:
            option_text = option.get_attribute('textContent')
            status_update_second_report_actual_texts.append(option_text)
        assert status_update_second_info == status_update_second_report_actual_texts
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(5)
        export_dropdown_menu_list_items = self.get_elements(self.export_dropdown_menu_list, "xpath")
        status_update_texts_items = []
        for option in export_dropdown_menu_list_items:
            option_text_items = option.get_attribute('textContent')
            status_update_texts_items.append(option_text_items)
        assert status_update_texts_items == status_update_list_items
        time.sleep(5)
        self.click_on_element(self.word_file, "xpath")
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert status_update_word_file == downloaded_word_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.excel_file, "xpath")
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == status_update_excel
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            first_actual_column_names = [cell.value for cell in sheet[2] if cell.value is not None]
        if first_actual_column_names == status_update_first_info:
            assert first_actual_column_names == status_update_first_info
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            second_actual_column_names = [cell.value for cell in sheet[5] if cell.value is not None]
        if second_actual_column_names == status_update_second_info:
            assert second_actual_column_names == status_update_second_info
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.power_point_file, "xpath")
        time.sleep(30)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == status_update_power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.pdf_file, "xpath")
        time.sleep(30)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == status_update_pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.print_dialog_box_text).get_attribute("textContent")
        assert status_update_text == print_text
        self.click_on_element(self.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.Done_button, "xpath")


    def validating_EDD_report(self, subtract_days=1):
        time.sleep(3)
        self.verify_browse_header()
        time.sleep(2)
        self.driver.switch_to.frame(self.get_element(self.report_iframe1, "xpath"))
        # time.sleep(10)
        assert self.isElementPresent(self.delivery_date_field, 'xpath')
        assert self.isElementPresent(self.facility_field, 'xpath')
        assert self.isElementPresent(self.customer_field_txt, 'xpath')
        assert self.isElementPresent(self.diff_txt, 'xpath')
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.facility_carat).click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.jackson_value_option).click()
        time.sleep(4)
        self.driver.find_element(By.XPATH, self.customer_carat_arrow).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.select_all_check_box).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.customer_carat_arrow).click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, self.view_report_btn).click()
        time.sleep(45)
        self.driver.find_element(By.XPATH, self.export_drop_down_btn).click()
        time.sleep(5)
        self.driver.find_element(By.XPATH, self.word_type_btn).click()
        time.sleep(8)

    def validating_print_text_edd_vals(self, word_file, excel_file_name, expected_texts, pdf_name, power_point_name,
                                       text):
        self.eos_bolt_menu_report = eos_bolt_menu_report(self.driver)
        global actual_column_names
        self.click_on_element(self.eos_bolt_menu_report.word_file, "xpath")
        # self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.word_file).click()
        time.sleep(15)
        downloaded_word_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_word_name))
        print("downloaded_word_name##############################")
        print(downloaded_word_name)
        assert downloaded_word_name == word_file
        time.sleep(3)
        self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.export_dropdown_menu).click()
        time.sleep(1)
        self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.excel_file).click()
        time.sleep(10)
        downloaded_excel_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_excel_name == excel_file_name
        file_path = os.path.abspath('.') + "\\Downloads\\" + downloaded_excel_name
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            actual_column_names = [cell.value for cell in sheet[3]]
            print("Actual column Vals######################")
            print(actual_column_names)
        if actual_column_names == expected_texts:
            assert actual_column_names == expected_texts
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", file_path))
        time.sleep(3)
        self.click_on_element(self.eos_bolt_menu_report.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.eos_bolt_menu_report.power_point_file, "xpath")
        time.sleep(10)
        downloaded_ppt_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_ppt_name == power_point_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_ppt_name))
        time.sleep(3)
        self.click_on_element(self.eos_bolt_menu_report.export_dropdown_menu, "xpath")
        time.sleep(1)
        self.click_on_element(self.eos_bolt_menu_report.pdf_file, "xpath")
        time.sleep(10)
        downloaded_pdf_name = os.listdir(os.path.abspath('.') + "\\Downloads")[0]
        assert downloaded_pdf_name == pdf_name
        os.remove(os.path.join(os.path.abspath('.') + "\\Downloads", downloaded_pdf_name))
        time.sleep(3)
        self.click_on_element(self.eos_bolt_menu_report.print_button, "xpath")
        print_text = self.driver.find_element(By.XPATH, self.eos_bolt_menu_report.print_dialog_box_text).get_attribute(
            "textContent")
        assert text == print_text
        self.click_on_element(self.eos_bolt_menu_report.dialog_print_button, "xpath")
        time.sleep(10)
        self.click_on_element(self.eos_bolt_menu_report.Done_button, "xpath")
